import os
import subprocess
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def docx_to_pdf(docx_path, output_dir):
    """Uses LibreOffice headless to convert a docx to PDF. Returns the pdf path,
    or None if LibreOffice ('soffice') is not available on this machine."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None
    subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", output_dir, docx_path],
        check=True, capture_output=True, timeout=60,
    )
    base = os.path.splitext(os.path.basename(docx_path))[0]
    pdf_path = os.path.join(output_dir, base + ".pdf")
    return pdf_path if os.path.exists(pdf_path) else None


def export_approval_note_excel(note, shipment, invoices, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Approval Note"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    ws["A1"] = note.document_number
    ws["A1"].font = bold
    ws["F1"] = note.document_date.strftime("%d/%m/%Y")
    ws["F1"].font = bold

    ws["A3"] = "Subject"
    ws["A3"].font = bold
    ws.merge_cells("B3:F3")
    ws["B3"] = note.subject_text

    ws["A5"] = "Para 1"
    ws["A5"].font = bold
    ws.merge_cells("B5:F5")
    ws["B5"] = note.paragraph_1

    headers = ["Sl No", "Consignment Charges", "Invoice/Ref", "Amount", "CHA Ref", "SYS Ref"]
    header_row = 7
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = bold
        c.alignment = center
        c.fill = fill

    row = header_row + 1
    total = 0
    for idx, inv in enumerate(invoices, start=1):
        vendor = f" ({inv.vendor_name})" if inv.vendor_name else ""
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f"{inv.short_label()}{vendor}")
        ws.cell(row=row, column=3, value=inv.invoice_number or "")
        amt = float(inv.invoice_amount or 0)
        total += amt
        ws.cell(row=row, column=4, value=amt).number_format = '"₹"#,##0.00'
        ws.cell(row=row, column=5, value=inv.cha_ref or "")
        ws.cell(row=row, column=6, value=inv.srn or "")
        row += 1

    ws.cell(row=row, column=3, value="Total").font = bold
    tc = ws.cell(row=row, column=4, value=total)
    tc.font = bold
    tc.number_format = '"₹"#,##0.00'

    row += 2
    if note.paragraph_2:
        ws.cell(row=row, column=1, value="Para 2").font = bold
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=note.paragraph_2)
        row += 2

    if note.paragraph_3:
        ws.cell(row=row, column=1, value="Para 3").font = bold
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=note.paragraph_3)
        row += 2

    ws.cell(row=row, column=1, value="Para 4").font = bold
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value=note.paragraph_4)

    widths = [8, 30, 22, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def export_register_excel(rows, headers, title, output_path):
    """Generic register/report exporter (Shipment Register, Invoice Register, etc.)"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    bold = Font(bold=True)
    fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
